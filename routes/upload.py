from fastapi import APIRouter, HTTPException, UploadFile, File
from typing import List
import pandas as pd
import io
import csv
import sys, os
import requests as _requests
from openpyxl import load_workbook
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from .store_utils import get_shopify_client

router = APIRouter()
PREVIEW_ROWS = 10

def get_existing_products():
    """Fetch ALL existing products from Shopify to check for duplicates.
    
    This is CRITICAL for duplicate detection. Raises exception if fetch fails.
    """
    existing = {"titles": set(), "skus": set(), "handles": set(), "product_ids": set()}
    
    try:
        client = get_shopify_client()
        print("🔍 [CRITICAL] Fetching ALL existing Shopify products for duplicate check...")
        
        try:
            result = client.get_products(fetch_all=True)
        except _requests.exceptions.HTTPError as http_err:
            # Token expired mid-fetch — refresh and retry
            if http_err.response is not None and http_err.response.status_code == 401:
                print("🔄 [CRITICAL] Token expired while fetching existing products, refreshing...")
                client = get_shopify_client()
                result = client.get_products(fetch_all=True)
            else:
                raise
        
        products = result.get("products", [])
        if not products:
            print("⚠️ [WARNING] No products found in Shopify! Continuing with empty existing products set.")
            print("📦 Existing products: 0 titles, 0 SKUs, 0 handles")
            return existing
        
        for product in products:
            product_id = product.get("id")
            if product_id:
                existing["product_ids"].add(str(product_id))
            
            title = (product.get("title") or "").lower().strip()
            if title:
                existing["titles"].add(title)

            handle = (product.get("handle") or "").lower().strip()
            if handle:
                existing["handles"].add(handle)

            for variant in product.get("variants", []):
                sku = (variant.get("sku") or "").lower().strip()
                if sku:
                    existing["skus"].add(sku)

        print(f"✅ SUCCESS: Fetched {len(products)} existing products")
        print(f"📦 Existing products: {len(existing['titles'])} titles, {len(existing['skus'])} SKUs, {len(existing['handles'])} handles")
        return existing

    except Exception as e:
        print(f"❌ [CRITICAL ERROR] Failed to fetch existing products: {e}")
        print(f"❌ CANNOT PROCEED WITH UPLOAD - DUPLICATE CHECK FAILED!")
        raise HTTPException(
            status_code=500, 
            detail=f"Critical error: Could not fetch existing products from Shopify. Duplicate check failed. Error: {str(e)}"
        )


def parse_file(content, filename, preview=False):
    if filename.endswith(".csv"):
        if preview:
            return pd.read_csv(io.BytesIO(content), nrows=PREVIEW_ROWS)
        return pd.read_csv(io.BytesIO(content))
    elif filename.endswith((".xlsx", ".xls")):
        if preview:
            return pd.read_excel(io.BytesIO(content), nrows=PREVIEW_ROWS)
        return pd.read_excel(io.BytesIO(content))
    raise ValueError("Unsupported file type. Use .csv, .xlsx or .xls")


def estimate_total_rows(content, filename):
    if filename.endswith(".csv"):
        reader = csv.reader(io.StringIO(content.decode("utf-8-sig", errors="replace")))
        line_count = sum(1 for _ in reader)
        return max(line_count - 1, 0)

    if filename.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            return max((worksheet.max_row or 1) - 1, 0)
        finally:
            workbook.close()

    # Legacy .xls fallback keeps previous behavior.
    if filename.endswith(".xls"):
        return len(pd.read_excel(io.BytesIO(content)))

    raise ValueError("Unsupported file type. Use .csv, .xlsx or .xls")


@router.post("/preview")
async def preview_file(file: UploadFile = File(...)):
    try:
        content = await file.read()
        total_rows = estimate_total_rows(content, file.filename)
        df = parse_file(content, file.filename, preview=True)
        df = df.fillna("")
        return {
            "filename": file.filename,
            "total_rows": total_rows,
            "columns": list(df.columns),
            "preview": df.head(PREVIEW_ROWS).to_dict(orient="records")
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/parse")
async def parse_full_file(file: UploadFile = File(...)):
    try:
        content = await file.read()
        df = parse_file(content, file.filename)
        df = df.fillna("")
        return {
            "filename": file.filename,
            "total_rows": len(df),
            "columns": list(df.columns),
            "data": df.to_dict(orient="records")
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/validate")
async def validate_products(products: List[dict]):
    """
    Validate products before upload:
    - Checks for required fields
    - Detects duplicates in uploaded file
    - Detects duplicates against existing Shopify products
    """
    errors = []
    valid_count = 0
    duplicates = []
    existing_duplicates = []

    try:
        existing_products = get_existing_products()
    except HTTPException:
        # get_existing_products() will raise HTTPException if it fails
        # Let it bubble up so frontend sees the error
        raise
    
    seen_titles = set()
    seen_skus = set()
    seen_handles = set()

    for i, product in enumerate(products):
        title = (product.get("title") or product.get("Title") or "").strip()
        sku = (product.get("sku") or product.get("Variant SKU") or "").strip()
        handle = (product.get("handle") or product.get("Handle") or "").strip()

        has_data = any(
            str(v).strip() for k, v in product.items()
            if k.lower() not in ["handle", "published", "option1 name", "option1 value"]
            and str(v).strip()
        )

        if not title and has_data:
            errors.append({
                "row": i + 1,
                "field": "title",
                "message": f"Row {i + 1} has data but is missing a Title"
            })
        elif title:
            title_lower = (title or "").lower().strip()
            sku_lower = (sku or "").lower().strip() if sku else ""
            handle_lower = (handle or "").lower().strip() if handle else ""

            # Check for duplicate in uploaded file
            if title_lower in seen_titles:
                duplicates.append({"row": i + 1, "title": title, "reason": "Duplicate title in uploaded file"})
            # Check for title duplicate in Shopify
            elif title_lower in existing_products["titles"]:
                existing_duplicates.append({"row": i + 1, "title": title, "reason": "Title already exists in Shopify"})
            # Check for SKU duplicate in Shopify
            elif sku_lower and sku_lower in existing_products["skus"]:
                existing_duplicates.append({"row": i + 1, "title": title, "sku": sku, "reason": "SKU already exists in Shopify"})
            # Check for handle duplicate in Shopify
            elif handle_lower and handle_lower in existing_products["handles"]:
                existing_duplicates.append({"row": i + 1, "title": title, "handle": handle, "reason": "Handle already exists in Shopify"})
            # Check for duplicate handle in uploaded file
            elif handle_lower and handle_lower in seen_handles:
                duplicates.append({"row": i + 1, "title": title, "handle": handle, "reason": "Duplicate handle in uploaded file"})
            else:
                valid_count += 1

            seen_titles.add(title_lower)
            if sku_lower:
                seen_skus.add(sku_lower)
            if handle_lower:
                seen_handles.add(handle_lower)

    all_issues = errors + duplicates + existing_duplicates
    return {
        "valid": len(all_issues) == 0,
        "error_count": len(errors),
        "duplicate_count": len(duplicates),
        "existing_duplicate_count": len(existing_duplicates),
        "errors": errors,
        "duplicates": duplicates,
        "existing_duplicates": existing_duplicates,
        "total": len(products),
        "valid_products": valid_count,
        "summary": {
            "will_be_uploaded": valid_count,
            "will_be_skipped": len(duplicates) + len(existing_duplicates),
            "has_errors": len(errors) > 0
        }
    }


@router.post("/push-to-shopify")
async def push_to_shopify(products: List[dict]):
    """
    Upload products to Shopify with comprehensive duplicate checking.
    
    This endpoint:
    1. Fetches all existing Shopify products (CRITICAL)
    2. Filters out any duplicates by title, SKU, or handle
    3. Groups multi-variant products
    4. Creates only new products
    """
    try:
        client = get_shopify_client()
        results = {
            "success": [],
            "errors": [],
            "skipped": [],
            "total": 0,
            "created": 0,
            "skipped_count": 0
        }

        # STEP 1: Fetch FRESH existing products before every push
        print("\n" + "="*80)
        print("🔍 STEP 1: Fetching existing Shopify products...")
        print("="*80)
        try:
            existing_products = get_existing_products()
        except HTTPException as e:
            # Critical error - can't proceed without knowing what exists
            print(f"❌ CRITICAL: Cannot continue without duplicate check: {e.detail}")
            raise
        
        print(f"✅ STEP 1 Complete: Will check all uploads against {len(existing_products['titles'])} existing titles\n")

        seen_titles = set()      # track titles within this upload batch
        seen_skus = set()        # track SKUs within this upload batch
        seen_handles = set()     # track handles within this upload batch

        # STEP 2: Group rows by Handle for multi-variant products
        print("="*80)
        print("🔍 STEP 2: Processing and grouping products...")
        print("="*80)
        
        grouped = {}
        skipped_rows = []
        
        for row_idx, row in enumerate(products):
            handle = (row.get("Handle") or row.get("handle") or row.get("HANDLE") or "").strip()
            title = (row.get("Title") or row.get("title") or row.get("TITLE") or "").strip()
            sku = (row.get("Variant SKU") or row.get("sku") or row.get("VARIANT SKU") or "").strip()

            if not handle and not title:
                continue

            key = handle or title
            title_lower = (title or "").lower().strip()
            sku_lower = (sku or "").lower().strip() if sku else ""
            handle_lower = (handle or "").lower().strip() if handle else ""

            skip_reason = None

            # ✋ DUPLICATE CHECK 0: SKU already exists in Shopify
            if sku_lower:
                if sku_lower in existing_products["skus"]:
                    skip_reason = f"SKU '{sku}' already exists in Shopify"
                elif sku_lower in seen_skus:
                    skip_reason = f"SKU '{sku}' is duplicate within uploaded file"

            # ✋ DUPLICATE CHECK 1: Product title already exists in Shopify
            if not skip_reason and title_lower:
                if title_lower in existing_products["titles"]:
                    skip_reason = f"Title '{title}' already exists in Shopify"
                elif title_lower in seen_titles:
                    # This is a variant row for an already-seen product, add it
                    if key in grouped:
                        existing_variant_skus = [v.get("sku", "").lower().strip() for v in grouped[key]["variants"]]
                        if sku_lower and sku_lower not in existing_variant_skus:
                            _add_variant_to_group(grouped[key], row)
                            if sku_lower:
                                seen_skus.add(sku_lower)
                    continue

            # ✋ DUPLICATE CHECK 2: Handle already exists in Shopify
            if not skip_reason and handle_lower:
                if handle_lower in existing_products["handles"]:
                    skip_reason = f"Handle '{handle}' already exists in Shopify"
                elif handle_lower in seen_handles:
                    skip_reason = f"Handle '{handle}' is duplicate within uploaded file"

            # If we should skip this row, log it
            if skip_reason:
                print(f"⏭️  Row {row_idx + 1}: SKIPPED — {skip_reason}")
                results["skipped"].append({
                    "row": row_idx + 1,
                    "title": title or "(no title)",
                    "reason": skip_reason
                })
                results["skipped_count"] += 1
                skipped_rows.append(row)
                continue

            # ✅ This is a new product — add to group
            if title_lower:
                seen_titles.add(title_lower)
            if sku_lower:
                seen_skus.add(sku_lower)
            if handle_lower:
                seen_handles.add(handle_lower)

            if key not in grouped:
                status = (row.get("Status") or row.get("status") or row.get("STATUS") or "").strip().lower()
                if status not in ["active", "draft", "archived"]:
                    status = "active"
                
                grouped[key] = {
                    "title": title,
                    "body_html": row.get("Body (HTML)") or row.get("body_html") or row.get("BODY (HTML)") or "",
                    "vendor": row.get("Vendor") or row.get("vendor") or row.get("VENDOR") or "",
                    "product_type": row.get("Type") or row.get("product_type") or row.get("TYPE") or "",
                    "tags": row.get("Tags") or row.get("tags") or row.get("TAGS") or "",
                    "status": status,
                    "variants": [],
                    "images": [],
                    "seo_title": "",
                    "seo_description": ""
                }
                
                grouped[key]["seo_title"] = row.get("SEO TITLE") or row.get("Seo Title") or row.get("SEO Title") or row.get("seo_title") or ""
                grouped[key]["seo_description"] = row.get("SEO DESCRIPTION") or row.get("Seo Description") or row.get("SEO Description") or row.get("seo_description") or ""

            _add_variant_to_group(grouped[key], row)

        print(f"✅ STEP 2 Complete: {len(grouped)} new products to create, {results['skipped_count']} will be skipped\n")
        results["total"] = len(grouped) + results["skipped_count"]

        # STEP 3: Create products in Shopify
        print("="*80)
        print(f"🚀 STEP 3: Uploading {len(grouped)} new products...")
        print("="*80)
        
        for product_idx, (key, product_data) in enumerate(grouped.items(), 1):
            try:
                seo_title = product_data.pop("seo_title", "")
                seo_desc = product_data.pop("seo_description", "")
                
                if not product_data["variants"]:
                    del product_data["variants"]
                if not product_data["images"]:
                    del product_data["images"]

                title = product_data.get("title")
                print(f"  [{product_idx}/{len(grouped)}] 🚀 Creating: {title}")
                
                try:
                    r = client.create_product(product_data)
                except _requests.exceptions.HTTPError as http_err:
                    if http_err.response is not None and http_err.response.status_code == 401:
                        print(f"      🔄 Token expired, refreshing...")
                        client = get_shopify_client()
                        r = client.create_product(product_data)
                    else:
                        raise
                
                product_id = r.get("product", {}).get("id")
                
                if product_id and (seo_title or seo_desc):
                    try:
                        print(f"      📝 Setting SEO...")
                        client.set_product_seo(product_id, title=seo_title or None, description=seo_desc or None)
                    except Exception as seo_err:
                        print(f"      ⚠️  SEO failed (non-blocking): {seo_err}")
                
                results["success"].append({
                    "title": title,
                    "id": product_id
                })
                results["created"] += 1
                print(f"      ✅ Created successfully (ID: {product_id})")
                
            except Exception as e:
                error_msg = str(e)
                print(f"      ❌ FAILED: {error_msg}")
                results["errors"].append({
                    "title": product_data.get("title"),
                    "error": error_msg
                })

        print("\n" + "="*80)
        print(f"📊 FINAL RESULTS:")
        print(f"  ✅ Created:  {results['created']}")
        print(f"  ⏭️  Skipped: {results['skipped_count']}")
        print(f"  ❌ Errors:   {len(results['errors'])}")
        print(f"  📦 Total:    {results['total']}")
        print("="*80 + "\n")
        
        return results

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ FATAL ERROR: {e}")
        raise HTTPException(status_code=400, detail=str(e))


def _add_variant_to_group(product_group: dict, row: dict):
    """Extract variant data from a row and add to product group"""
    variant = {}
    price = row.get("Variant Price") or row.get("price") or row.get("VARIANT PRICE") or ""
    sku = row.get("Variant SKU") or row.get("sku") or row.get("VARIANT SKU") or ""
    qty = row.get("Variant Inventory Qty") or row.get("inventory_quantity") or row.get("VARIANT INVENTORY QTY") or ""
    barcode = row.get("Variant Barcode") or row.get("barcode") or row.get("VARIANT BARCODE") or ""
    compare_price = row.get("Variant Compare At Price") or row.get("compare_at_price") or row.get("VARIANT COMPARE AT PRICE") or ""

    if price:
        variant["price"] = str(price)
    if sku:
        variant["sku"] = str(sku)
    if qty:
        try:
            variant["inventory_quantity"] = int(float(str(qty)))
        except:
            pass
    if barcode:
        variant["barcode"] = str(barcode)
    if compare_price:
        variant["compare_at_price"] = str(compare_price)

    if variant:
        product_group["variants"].append(variant)

    img = row.get("Image Src") or row.get("image_src") or row.get("IMAGE URLS") or row.get("Image URLs") or row.get("IMAGE SRC") or ""
    if img:
        img_src = str(img)
        existing_srcs = [i["src"] for i in product_group["images"]]
        if img_src not in existing_srcs:
            product_group["images"].append({"src": img_src})
